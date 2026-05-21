"""
Fetch Citadele Bank App Store reviews from Estonia, Lithuania and Latvia
storefronts using Apple's public Customer Reviews RSS feed.

Output: reviews.json and reviews.xlsx in the current directory.

Apple limits this feed to roughly the most recent ~500 reviews per country
(10 pages * 50 entries). For older reviews you'd need App Store Connect API
access (requires being the app owner with an API key).
"""

from __future__ import annotations

import json
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_ID = "495139240"  # Citadele Bank
COUNTRIES = {
    "ee": "Estonia",
    "lt": "Lithuania",
    "lv": "Latvia",
}
MAX_PAGES = 10  # Apple caps the feed around page 10
PAGE_SIZE_HINT = 50
REQUEST_TIMEOUT = 20
THROTTLE_SECONDS = 1.0  # be polite; Apple throttles aggressive callers

URL_TEMPLATE = (
    "https://itunes.apple.com/{country}/rss/customerreviews/"
    "page={page}/id={app_id}/sortBy=mostRecent/json"
)

HEADERS = {
    "User-Agent": (
        "CitadeleReviewFetcher/1.0 (+internal tooling; contact: yourname@citadele.lv)"
    )
}


@dataclass
class Review:
    country_code: str
    country_name: str
    review_id: str
    rating: int
    title: str
    content: str
    author_name: str
    author_uri: str
    app_version: str
    updated: str  # ISO timestamp from Apple
    page: int


def _safe(d: dict, *keys, default: str = "") -> str:
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur if isinstance(cur, str) else default


def fetch_page(country: str, page: int) -> list[dict]:
    url = URL_TEMPLATE.format(country=country, page=page, app_id=APP_ID)
    resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    if resp.status_code == 403:
        print(f"  [!] 403 from Apple on {country} page {page} — being throttled, stopping this country.")
        return []
    resp.raise_for_status()
    data = resp.json()
    entries = data.get("feed", {}).get("entry", []) or []
    # On page 1 the very first entry is usually app metadata, not a review.
    # Filter to entries that actually have im:rating.
    return [e for e in entries if isinstance(e, dict) and "im:rating" in e]


def parse_entry(entry: dict, country_code: str, country_name: str, page: int) -> Review:
    return Review(
        country_code=country_code,
        country_name=country_name,
        review_id=_safe(entry, "id", "label"),
        rating=int(_safe(entry, "im:rating", "label", default="0") or 0),
        title=_safe(entry, "title", "label"),
        content=_safe(entry, "content", "label"),
        author_name=_safe(entry, "author", "name", "label"),
        author_uri=_safe(entry, "author", "uri", "label"),
        app_version=_safe(entry, "im:version", "label"),
        updated=_safe(entry, "updated", "label"),
        page=page,
    )


def fetch_country(country_code: str, country_name: str) -> list[Review]:
    print(f"\n== {country_name} ({country_code}) ==")
    collected: dict[str, Review] = {}  # dedupe by review_id
    for page in range(1, MAX_PAGES + 1):
        try:
            raw = fetch_page(country_code, page)
        except requests.RequestException as exc:
            print(f"  [!] Network error on page {page}: {exc}. Stopping this country.")
            break
        if not raw:
            print(f"  page {page}: no more reviews, stopping.")
            break
        parsed = [parse_entry(e, country_code, country_name, page) for e in raw]
        new_count = sum(1 for r in parsed if r.review_id not in collected)
        for r in parsed:
            collected[r.review_id] = r
        print(f"  page {page}: {len(parsed)} entries ({new_count} new). total so far: {len(collected)}")
        if new_count == 0:
            print("  page returned only duplicates — assuming end of feed.")
            break
        time.sleep(THROTTLE_SECONDS)
    return list(collected.values())


def write_json(reviews: list[Review], path: Path) -> None:
    payload = {
        "app_id": APP_ID,
        "app_name": "Citadele Bank",
        "fetched_at_utc": datetime.now(timezone.utc).isoformat(),
        "review_count": len(reviews),
        "countries": list(COUNTRIES.values()),
        "reviews": [asdict(r) for r in reviews],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {path} ({len(reviews)} reviews)")


def write_xlsx(reviews: list[Review], path: Path) -> None:
    wb = Workbook()

    # --- Sheet 1: all reviews ---
    ws = wb.active
    ws.title = "Reviews"
    headers = [
        "Country", "Country code", "Rating", "Title", "Content",
        "Author", "App version", "Updated (UTC)", "Review ID", "Author URI", "Page",
    ]
    ws.append(headers)
    for r in reviews:
        ws.append([
            r.country_name, r.country_code, r.rating, r.title, r.content,
            r.author_name, r.app_version, r.updated, r.review_id, r.author_uri, r.page,
        ])

    # Header style
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E78")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Column widths
    widths = [12, 6, 8, 40, 70, 22, 12, 22, 14, 40, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Wrap text in title/content and freeze header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in {"D", "E"})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: summary stats ---
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Country", "Reviews fetched", "Avg rating (formula)", "5★", "4★", "3★", "2★", "1★"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill

    countries_in_order = list(COUNTRIES.values())
    for i, country in enumerate(countries_in_order, start=2):
        ws2.cell(row=i, column=1, value=country)
        # Count reviews for this country from Reviews sheet
        country_range = "Reviews!A:A"
        rating_range = "Reviews!C:C"
        ws2.cell(row=i, column=2, value=f'=COUNTIF({country_range},A{i})')
        ws2.cell(row=i, column=3, value=f'=IFERROR(AVERAGEIF({country_range},A{i},{rating_range}),0)')
        ws2.cell(row=i, column=3).number_format = "0.00"
        for star, col in zip([5, 4, 3, 2, 1], [4, 5, 6, 7, 8]):
            ws2.cell(row=i, column=col,
                     value=f'=COUNTIFS({country_range},A{i},{rating_range},{star})')

    # Totals row
    total_row = len(countries_in_order) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    ws2.cell(row=total_row, column=3,
             value=f'=IFERROR(AVERAGE(Reviews!C2:C{len(reviews) + 1}),0)')
    ws2.cell(row=total_row, column=3).number_format = "0.00"
    for col in range(4, 9):
        col_letter = get_column_letter(col)
        ws2.cell(row=total_row, column=col,
                 value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")

    for c in ws2[total_row]:
        c.font = Font(bold=True)

    for i, w in enumerate([14, 18, 22, 8, 8, 8, 8, 8], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)
    print(f"Wrote {path}")


def main() -> int:
    all_reviews: list[Review] = []
    for code, name in COUNTRIES.items():
        all_reviews.extend(fetch_country(code, name))

    if not all_reviews:
        print("No reviews collected.")
        return 1

    out_dir = Path("output")
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")

    write_json(all_reviews, out_dir / f"citadele_reviews_{stamp}.json")
    write_xlsx(all_reviews, out_dir / f"citadele_reviews_{stamp}.xlsx")

    # Also overwrite a "latest" copy for easy automation
    write_json(all_reviews, out_dir / "citadele_reviews_latest.json")
    write_xlsx(all_reviews, out_dir / "citadele_reviews_latest.xlsx")

    print(f"\nDone. Total reviews: {len(all_reviews)}")
    by_country: dict[str, int] = {}
    for r in all_reviews:
        by_country[r.country_name] = by_country.get(r.country_name, 0) + 1
    for c, n in by_country.items():
        print(f"  {c}: {n}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
