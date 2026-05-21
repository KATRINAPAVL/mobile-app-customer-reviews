"""
Fetch Citadele Bank Google Play Store reviews via google-play-scraper.

Google Play returns a different curated subset of reviews depending on the
(country, language) the API is queried with. To catch the broadest possible
range of reviewers — including Russian and English speakers in the Baltics,
and international users — we do 9 locale passes covering native, Russian,
and English overlays for each of the three Baltic countries. Results are
deduplicated by review ID, and 'fetched_via' records every pass that
returned each review.

We also do a 2nd pass with Sort.MOST_RELEVANT because Google's
"most_relevant" algorithm sometimes surfaces reviews the "newest" sort
misses (older reviews from heavy users, multi-language reviews, etc).

Output (in ./output/):
  citadele_play_reviews_<timestamp>.json / .xlsx
  citadele_play_reviews_latest.json / .xlsx
"""

from __future__ import annotations

import json
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path

from google_play_scraper import Sort, reviews
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_ID = "lv.citadele.mobile"

# 9 locale passes: native + Russian + English overlay per country
# Each tuple: (gl country code, hl language code, display country label)
LOCALES = [
    ("lv", "lv", "Latvia"),
    ("lv", "ru", "Latvia"),
    ("lv", "en", "Latvia"),
    ("lt", "lt", "Lithuania"),
    ("lt", "ru", "Lithuania"),
    ("lt", "en", "Lithuania"),
    ("ee", "et", "Estonia"),
    ("ee", "ru", "Estonia"),
    ("ee", "en", "Estonia"),
]

# Both sort orders combined catch reviews the other misses
SORTS = [Sort.NEWEST, Sort.MOST_RELEVANT]

BATCH_SIZE = 200
MAX_PER_PASS = 2000
THROTTLE_SECONDS = 0.3


@dataclass
class Review:
    review_id: str
    primary_locale: str
    fetched_via: list[str] = field(default_factory=list)
    primary_country: str = ""
    rating: int = 0
    content: str = ""
    author_name: str = ""
    app_version: str = ""
    updated_iso: str = ""
    thumbs_up: int = 0
    reply_content: str = ""
    reply_at_iso: str = ""


def fetch_pass(country: str, lang: str, display: str, sort: Sort) -> list[dict]:
    sort_label = "newest" if sort == Sort.NEWEST else "most_relevant"
    label = f"{country}/{lang}/{sort_label}"
    print(f"\n== {display} ({label}) ==")
    collected: list[dict] = []
    continuation_token = None
    while len(collected) < MAX_PER_PASS:
        try:
            if continuation_token is None:
                batch, continuation_token = reviews(
                    APP_ID,
                    lang=lang,
                    country=country,
                    sort=sort,
                    count=BATCH_SIZE,
                )
            else:
                batch, continuation_token = reviews(
                    APP_ID,
                    continuation_token=continuation_token,
                )
        except Exception as exc:
            print(f"  [!] error: {exc}. Stopping this pass.")
            break

        if not batch:
            print("  no more reviews returned.")
            break

        collected.extend(batch)
        print(f"  batch: {len(batch)} (total this pass: {len(collected)})")

        if continuation_token is None:
            print("  no continuation token, end of feed.")
            break

        time.sleep(THROTTLE_SECONDS)

    return collected


def merge_reviews(per_pass: dict[str, list[dict]],
                  pass_to_country: dict[str, str]) -> list[Review]:
    """Dedupe by reviewId, keep first-found pass as primary, record all passes that found it."""
    merged: dict[str, Review] = {}

    for label, batch in per_pass.items():
        country_display = pass_to_country[label]
        for r in batch:
            rid = r.get("reviewId", "") or ""
            if not rid:
                continue
            if rid in merged:
                if label not in merged[rid].fetched_via:
                    merged[rid].fetched_via.append(label)
                continue

            at = r.get("at")
            replied_at = r.get("repliedAt")
            merged[rid] = Review(
                review_id=rid,
                primary_locale=label,
                fetched_via=[label],
                primary_country=country_display,
                rating=int(r.get("score") or 0),
                content=(r.get("content") or "").strip(),
                author_name=(r.get("userName") or "").strip(),
                app_version=(r.get("reviewCreatedVersion") or "") or "",
                updated_iso=at.isoformat() if at else "",
                thumbs_up=int(r.get("thumbsUpCount") or 0),
                reply_content=(r.get("replyContent") or "").strip(),
                reply_at_iso=replied_at.isoformat() if replied_at else "",
            )

    return list(merged.values())


def write_json(reviews_list: list[Review], path: Path) -> None:
    payload = {
        "app_id": APP_ID,
        "app_name": "Citadele Bank (Android)",
        "fetched_at_utc": datetime.now(timezone.utc).isoformat(),
        "review_count": len(reviews_list),
        "passes_used": [f"{g}/{h}" for (g, h, _) in LOCALES],
        "sorts_used": ["newest", "most_relevant"],
        "note": (
            "Google Play returns a different curated subset per locale. We do 9 "
            "(country, language) passes x 2 sort orders. 'fetched_via' lists every "
            "pass that returned the review. 'primary_country' = the first pass's "
            "country, used as a soft attribution — not a hard country fact, since "
            "Play does not expose reviewer geography."
        ),
        "reviews": [asdict(r) for r in reviews_list],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {path} ({len(reviews_list)} reviews)")


def write_xlsx(reviews_list: list[Review], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"
    headers = [
        "Primary country", "Primary locale", "Fetched via",
        "Rating", "Content", "Author", "App version",
        "Updated (UTC)", "Thumbs up", "Reply", "Reply at (UTC)", "Review ID",
    ]
    ws.append(headers)
    for r in reviews_list:
        ws.append([
            r.primary_country, r.primary_locale, ", ".join(r.fetched_via),
            r.rating, r.content, r.author_name, r.app_version,
            r.updated_iso, r.thumbs_up, r.reply_content, r.reply_at_iso, r.review_id,
        ])

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E78")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    widths = [16, 18, 24, 8, 70, 22, 12, 22, 10, 50, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in {"E", "J"})

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Primary country", "Reviews fetched", "Avg rating", "5★", "4★", "3★", "2★", "1★"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill

    countries = sorted({d for (_, _, d) in LOCALES})
    for i, country in enumerate(countries, start=2):
        ws2.cell(row=i, column=1, value=country)
        ws2.cell(row=i, column=2, value=f'=COUNTIF(Reviews!A:A,A{i})')
        ws2.cell(row=i, column=3, value=f'=IFERROR(AVERAGEIF(Reviews!A:A,A{i},Reviews!D:D),0)')
        ws2.cell(row=i, column=3).number_format = "0.00"
        for star, col in zip([5, 4, 3, 2, 1], [4, 5, 6, 7, 8]):
            ws2.cell(row=i, column=col,
                     value=f'=COUNTIFS(Reviews!A:A,A{i},Reviews!D:D,{star})')

    total_row = len(countries) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    ws2.cell(row=total_row, column=3,
             value=f'=IFERROR(AVERAGE(Reviews!D2:D{len(reviews_list) + 1}),0)')
    ws2.cell(row=total_row, column=3).number_format = "0.00"
    for col in range(4, 9):
        col_letter = get_column_letter(col)
        ws2.cell(row=total_row, column=col,
                 value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
    for c in ws2[total_row]:
        c.font = Font(bold=True)

    for i, w in enumerate([18, 18, 14, 8, 8, 8, 8, 8], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)
    print(f"Wrote {path}")


def main() -> int:
    per_pass: dict[str, list[dict]] = {}
    pass_to_country: dict[str, str] = {}

    for sort in SORTS:
        sort_label = "newest" if sort == Sort.NEWEST else "most_relevant"
        for country, lang, display in LOCALES:
            label = f"{country}/{lang}/{sort_label}"
            pass_to_country[label] = display
            per_pass[label] = fetch_pass(country, lang, display, sort)

    merged = merge_reviews(per_pass, pass_to_country)
    if not merged:
        print("No reviews collected.")
        return 1

    out_dir = Path("output")
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")

    write_json(merged, out_dir / f"citadele_play_reviews_{stamp}.json")
    write_xlsx(merged, out_dir / f"citadele_play_reviews_{stamp}.xlsx")
    write_json(merged, out_dir / "citadele_play_reviews_latest.json")
    write_xlsx(merged, out_dir / "citadele_play_reviews_latest.xlsx")

    print(f"\nDone. Unique reviews: {len(merged)}")
    by_country: dict[str, int] = {}
    for r in merged:
        by_country[r.primary_country] = by_country.get(r.primary_country, 0) + 1
    for c, n in by_country.items():
        print(f"  {c} (primary): {n}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
